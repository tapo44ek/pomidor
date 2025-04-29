
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from repository.database import get_db_connection


def save_views_to_excel(
    output_path,
    history_id,
    date=False,
):
    """РАБОЧИЙ ВАРИАНТ"""
    print('in func')
    try:
        views = ["rank", "new_apart_all", "res_of_rec", "where_not"]
        with get_db_connection() as conn:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for view in views:
                    print(f"Обработка представления: {view}")
                    try:
                        if view == "rank":
                            # Запросы для извлечения данных из базы данных
                            query_old_ranked = """
                                SELECT
                                    old_apart.affair_id as old_apart_id,
                                    old_apart.room_count,
                                    old_apart.living_area,
                                    old_apart.is_special_needs_marker,
                                    old_apart.full_living_area,
                                    old_apart.total_living_area,
                                    old_apart.district,
                                    old_apart.municipal_district,
                                    old_apart.house_address,
                                    old_apart.rank
                                FROM old_apart
                                WHERE (rsm_status <> 'снято' or rsm_status is null)
                            """

                            query_new_ranked = """
                                SELECT 
                                    new_apart.new_apart_id, 
                                    new_apart.room_count, 
                                    new_apart.living_area, 
                                    new_apart.for_special_needs_marker, 
                                    new_apart.full_living_area, 
                                    new_apart.total_living_area, 
                                    new_apart.district, 
                                    new_apart.municipal_district, 
                                    new_apart.house_address, 
                                    new_apart.rank
                                FROM new_apart 
                                WHERE rank is not NULL
                            """
                            print('all okay')
                            
                            # Подключаем фильтрацию для запросов

                            query_old_ranked += f" AND old_apart.history_id = {history_id}"

                            print('old_query', query_old_ranked)
                            query_new_ranked += f" AND new_apart.history_id = {history_id}"
                            print('new_query', query_old_ranked)
                            # Добавляем фильтрацию по дате, если указана
                            if date:
                                query_old_ranked += " AND old_apart.created_at = (SELECT MAX(created_at) FROM old_apart)"
                                query_new_ranked += " AND new_apart.created_at = (SELECT MAX(created_at) FROM new_apart)"


                            df_old_ranked = pd.read_sql(query_old_ranked, conn)
                            df_new_ranked = pd.read_sql(query_new_ranked, conn)
                            print(df_old_ranked)
                            print('----')
                            print(df_new_ranked)
                            # Получение максимальных рангов по количеству комнат из базы данных
                            max_rank_query = """
                                SELECT room_count, MAX(rank) as max_rank
                                FROM new_apart
                                GROUP BY room_count
                            """

                            max_rank_df = pd.read_sql(max_rank_query, conn)
                            print('max_rank_df', max_rank_df)

                            max_rank_by_room_count = max_rank_df.set_index("room_count")["max_rank"].to_dict()
                            print('max_rank_by_room_count', max_rank_by_room_count)

                            # Объединяем данные старых и новых квартир
                            df_combined = pd.concat(
                                [
                                    df_old_ranked.assign(status="old"),
                                    df_new_ranked.assign(status="new"),
                                ],
                                ignore_index=True,
                            )

                            # Присваиваем ранги и группируем данные
                            df_combined["Ранг"] = df_combined["rank"].astype(int)

                            df = (
                                df_combined.groupby(["room_count", "Ранг"])
                                .agg(
                                    Пот_ть=("old_apart_id", "count"),
                                    Ресурс=("new_apart_id", "count"),
                                )
                                .reset_index()
                            )

                            # Рассчитываем баланс
                            df["Баланс"] = df["Ресурс"] - df["Пот_ть"]

                            def add_totals(df, max_rank_by_room_count):
                                # Инициализируем итоговые переменные
                                total_potency = 0
                                total_resource = 0
                                total_balance = 0

                                # Список для хранения новых строк
                                new_rows = []
                                previous_row = None
                                start_rank = None

                                for i in range(len(df)):
                                    row = df.iloc[i].to_dict()
                                    current_rank = row["Ранг"]
                                    room_count = row["room_count"]
                                    max_rank = max_rank_by_room_count.get(room_count, 0) + 1

                                    if previous_row is not None:
                                        # Проверяем, является ли previous_row['Ранг'] строкой с диапазоном или целым числом
                                        if (
                                            isinstance(previous_row["Ранг"], str)
                                            and "-" in previous_row["Ранг"]
                                        ):
                                            previous_rank = int(
                                                previous_row["Ранг"].split("-")[-1]
                                            )
                                        else:
                                            previous_rank = previous_row["Ранг"]

                                        # Проверяем, можно ли объединять строки
                                        if (
                                            (previous_row["Ресурс"] == 0)
                                            and current_rank != max_rank
                                            and previous_rank != max_rank
                                        ):
                                            # Объединяем строки, если хотя бы одна из них имеет Пот_ть = 0
                                            previous_row["Пот_ть"] += row["Пот_ть"]
                                            previous_row["Ресурс"] += row["Ресурс"]
                                            previous_row["Баланс"] += row["Баланс"]
                                            # Обновляем диапазон рангов
                                            previous_row["Ранг"] = (
                                                f"{start_rank}-{current_rank}"
                                            )
                                        else:
                                            # Добавляем предыдущую строку в список новых строк, если она не None
                                            if previous_row is not None:
                                                new_rows.append(previous_row)

                                            # Устанавливаем новую строку как предыдущую
                                            previous_row = row
                                            start_rank = (
                                                current_rank  # Начало нового диапазона
                                            )
                                    else:
                                        # Устанавливаем первую строку как предыдущую
                                        previous_row = row
                                        start_rank = current_rank

                                # Не забываем добавить последнюю строку после цикла
                                if previous_row is not None:
                                    new_rows.append(previous_row)

                                # Преобразуем список в DataFrame
                                df_new = pd.DataFrame(new_rows)

                                # Пересчитываем итоговые значения
                                total_potency = df_new["Пот_ть"].sum()
                                total_resource = df_new["Ресурс"].sum()
                                total_balance = df_new["Баланс"].sum()

                                # Создаем строку с итогами
                                totals = pd.DataFrame(
                                    [
                                        {
                                            "Ранг": "Итог",
                                            "Пот_ть": total_potency,
                                            "Ресурс": total_resource,
                                            "Баланс": total_balance,
                                        }
                                    ]
                                )

                                # Добавляем строку с итогами в конец нового DataFrame
                                df_with_totals = pd.concat(
                                    [df_new, totals], ignore_index=True
                                )
                                
                                return df_with_totals

                            # Добавляем итоговые данные по рангу для каждого типа квартиры с использованием add_totals
                            result_data = []
                            for room in df["room_count"].unique():
                                room_df = df[df["room_count"] == room].copy()
                                grouped_df = add_totals(room_df, max_rank_by_room_count)
                                grouped_df["room_count"] = room
                                result_data.append(grouped_df)

                            # Объединяем результаты в одну таблицу
                            df_grouped = pd.concat(result_data, ignore_index=True)

                            if df_grouped.empty:
                                print(f"Представление {view} не вернуло данных.")
                            else:
                                # Начальная позиция для вставки данных на листе
                                sheet_name = "Ранг"
                                ws = writer.book.create_sheet(sheet_name)
                                current_row, current_col = 1, 1

                                # Настройка стилей
                                header_font = Font(bold=True)
                                header_fill = PatternFill(
                                    start_color="FFFF99",
                                    end_color="FFFF99",
                                    fill_type="solid",
                                )
                                header_alignment = Alignment(horizontal="center")

                                # Перебор всех типов комнат и запись данных в Excel
                                for room in df_grouped["room_count"].unique():
                                    room_df = df_grouped[df_grouped["room_count"] == room][
                                        ["Ранг", "Пот_ть", "Ресурс", "Баланс"]
                                    ]
                                    # Заголовок типа квартир
                                    ws.cell(
                                        row=current_row, column=current_col
                                    ).value = f"{room} комната(ы)"
                                    ws.cell(
                                        row=current_row, column=current_col
                                    ).font = header_font
                                    ws.cell(
                                        row=current_row, column=current_col
                                    ).alignment = header_alignment
                                    ws.merge_cells(
                                        start_row=current_row,
                                        start_column=current_col,
                                        end_row=current_row,
                                        end_column=current_col + 3,
                                    )
                                    current_row += 1

                                    # Запись заголовков
                                    for idx, col_name in enumerate(room_df.columns):
                                        cell = ws.cell(
                                            row=current_row, column=current_col + idx
                                        )
                                        cell.value = col_name
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = header_alignment

                                    current_row += 1

                                    # Запись данных
                                    for row in dataframe_to_rows(
                                        room_df, index=False, header=False
                                    ):
                                        for idx, value in enumerate(row):
                                            ws.cell(
                                                row=current_row, column=current_col + idx
                                            ).value = value
                                        current_row += 1

                                    # Переход к следующему блоку для следующего типа комнат
                                    current_row = 1
                                    current_col += len(room_df.columns) + 1

                        else:
                            query_params = []
                            query = f"SELECT * FROM public.{view}"

                            query += " WHERE 1=1 "
                            if view == "new_apart_all":
                                    query += f" AND history_id = ({history_id})"
                            elif view == 'where_not':
                                if history_id:
                                    query += f' AND history_id = ({history_id})'
                            else:
                                if history_id :
                                    query += f" AND id_истории_для_ресурса = {history_id}"
                                if history_id:
                                    query += f' AND id_истории_для_кпу = {history_id}'
                            print(f"Выполнение запроса для представления {view}")
                            try:
                                df = pd.read_sql(query, conn, params=query_params)
                                df = df.dropna(how="all")
                                df.to_excel(writer, sheet_name=view, index=False)
                            except Exception as e:
                                print(
                                    f"Ошибка выполнения запроса для представления {view}: {e}"
                                )
                    except Exception as e:
                        print(e)
    except Exception as e:
        print(f"Ошибка: {e}")